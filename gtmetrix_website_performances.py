'''Test Scenario: Check the website speed of all the website URLs listed in excel sheet. Results need to be shown in front of each website respectively with proper labeling clarifying the output.
Speed needs to be checked on -
i) URL: https://gtmetrix.com/ (For this, both grade and time need to be captured)

Also, mark the grades/time with red if:
i) grade is other than 'A' or 'B'.
ii) time is greater than 4 seconds'''

import pandas as pd
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
import openpyxl
from openpyxl.styles import Font, colors


class WebsitePerformanceCalculator:

    def read_excel(self):
        df = pd.read_excel("URL_list.xlsx")
        df.drop(df.columns[0], axis=1, inplace=True)
        header = df.iloc[1]
        new_df = pd.DataFrame(df.values[2:], columns=header)
        return new_df

    def write_excel(self, df, path=""):
        if path:
            filename = path + "/GtMetrix_website_performances.xlsx"
        else:
            filename = "GtMetrix_website_performances.xlsx"
        df.to_excel(filename, index=False, header=True)
        self.result_update(filename)

    def gtmetrix(self, driver, df, wait):
        gt_output_time = []
        gt_output_grade = []
        driver.maximize_window()
        driver.get("https://gtmetrix.com/")
        driver.implicitly_wait(3)

        for index, row in df.iterrows():
            data = row['Website URL']
            if pd.notnull(data):
                try:
                    tab = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='url']")))
                    submit = driver.find_element(By.XPATH, "//div[@class='analyze-form-button']//button[@type='submit']")
                    tab.clear()
                    time.sleep(2)
                    tab.send_keys(data)
                    submit.click()
                    elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='report-score']")))
                    grade_element = elements[0].find_element(By.XPATH, ".//i[contains(@class, 'grade')]")
                    grade = grade_element.get_attribute('class')
                    gt_output_grade.append(grade[-1])
                    time_element = driver.find_element(By.XPATH,
                                                       "//div[@class='report-page-detail'][1]//span[@class='report-page-detail-value']")
                    time_value = time_element.text
                    gt_output_time.append(time_value)
                    driver.back()
                    time.sleep(2)
                except TimeoutException:
                    gt_output_time.append("ERROR")
                    gt_output_grade.append("ERROR")
                    driver.back()
            else:
                gt_output_time.append("0")
                gt_output_grade.append("0")
        df['gt_metrix_grade'] = gt_output_grade
        df['gt_metrix_time'] = gt_output_time
        self.write_excel(df)

    def result_update(self, filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        max_row = ws.max_row
        ft = Font(color=colors.RED)
        valid_grade = ['A', 'B']
        values = ["ERROR", "0"]
        for row in range(2, (max_row + 1)):
            cell2 = "B{}".format(row)
            cell3 = "C{}".format(row)
            wc2 = ws.cell(row, 2)
            wc3 = ws.cell(row, 3)
            grade = ws[cell2].value
            time1 = ws[cell3].value
            if grade in values:
                wc2.font = ft
                wc3.font = ft
            else:
                if grade not in valid_grade:
                    wc2.font = ft
                if "." in time1:
                    new_time = float(time1[0:-1])
                    if new_time > 4.0:
                        wc3.font = ft
        wb.save(filename)


sa = WebsitePerformanceCalculator()
driverLocation = "C:\\Pranjul\\webdriver\\chromedriver.exe"  # as per your webderiver location
os.environ['webdriver.driver.chrome'] = driverLocation
driver = webdriver.Chrome(driverLocation)
wait = WebDriverWait(driver, 120, poll_frequency=2)
df = sa.read_excel()
sa.gtmetrix(driver, df, wait)